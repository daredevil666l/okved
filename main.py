import requests
import json
import time
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


class APIKeyManager:
    """Класс для управления API ключами с автоматической ротацией"""

    def __init__(self, api_keys):
        self.api_keys = api_keys
        self.current_index = 0
        self.lock = threading.Lock()
        self.active_keys_count = len(api_keys)

    def get_current_key(self):
        with self.lock:
            return self.api_keys[self.current_index]

    def rotate_key(self):
        with self.lock:
            self.current_index = (self.current_index + 1) % len(self.api_keys)
            return self.api_keys[self.current_index]

    def has_more_keys(self):
        return self.active_keys_count > 0


class NoticeApp:
    """Основной класс приложения"""

    def __init__(self, root):
        self.root = root
        self.root.title("Notice - Поиск организаций по ОКВЭД")
        self.root.geometry("1000x700")

        # Предустановленные API ключи
        self.api_keys = [
            'iYlIHiaHEgrURRR1',
            'MKXMk0GLaUN48g63', '8zg1SLEpQPe7DCDH', 'K2TX2471T12E4qMP',
            '8L1FghahT7pWWv8H', 'WgSDVbbjfakK4Pj2', 'Nx6JYT5tXlCDrT89',
            'yIvRc7Z29tZ2pvmw', 'o7rA5pXvQvUBmkCT', 'UYtyJAYCFuCIitpu',
            'ijo2nJdF2ZVMuyY8', 'bfcsAIOe0dS9bbeK', 'HZa0jsZSxUZI6x44',
            'wros2k7udZCInXbh', 'DoBA8V5nSXidszTO', 'iBtUfMDQH3klPsdx',
            'wWs0zpbtMAxAZoV1', 'TFOp6wBsqJxQD6sW', 'nvmRggGoc6av1XFn',
            'gC7Vinvt74myvtB2', 'xmHSJECg4PnW35XR', 'ywvmtFLGaVEIcI46','KFtiJ6TDH6KBnavp','j5tKSAJUDIX09HRL',
            'XZ20PTBKDDbdndvg','9dP6hmMdU3JpiNxb','7XzHUND4gVknLBPB',
            'KotL6Dd5oVevknZg','oNTHcW81rdgZTivx','uBmvp05s3J4yXkMD',
            'md5yGbllHARPLGSp','HN420ioHDMeGvzSr','FHdAVMiwJh8wth3w',
            'vRShEd3bBfc6lQoN','PKDlTpW61Y9oHhaD','DLdYuz2CPeMShrSL',
            'bFmJryzxKZithco4','RFOVm9SzD8baEHW9','J1f5iRuI5eQDEcJZ',
            '3gF3WaQn38kYKfIo','wOKDpjmXeukuqxbR','JdglNlgl5bfsYWGu',
            '3pGv0Mv6ODJK3jKS','59OG5yIbfhYAXsTb', '2W8ZVXfEBL6jNuV7',
            'MLOXKT9HPlPtalYz','dV1kGsDjYTR1CPyg','bwBrbqEvN96QydMY',
            'ioDOXSZmeG779XnS','0VCGdmu1wOC9BqQC','v3ZkRNmseYhkZbB7','17uHqcFh6QDvMeYs',
            '1K9jKKe08ngZE4Rc','eJtTICJQMD8j23M1','TJzJi7IOJhPpBtex',
            'TtCFNZBSsOIG3T3T','zthd0zOzJXP6RdcY','MeDnCkfcoGEkAkKd',
            'p8mIxSkwaZbaoy2Q','PORfSGvytK8irk2q','NHoqb8WihWAxddXc',
            'tCD7u7UrvSE5uUzr','PDFRDVqNsCetBHbV','LXvApBRjtbMHyBmq',
            'yeBdZPdpXVD4MI0O'


        ]

        self.setup_ui()

        # Флаги для контроля выполнения
        self.search_running = False
        self.animation_running = False
        self.animation_objects = []

        # Для хранения результатов
        self.collected_data = []

    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        # Верхняя панель с заголовком
        header_frame = tk.Frame(self.root, bg="#4b7bec", height=60)
        header_frame.pack(fill=tk.X)

        title_label = tk.Label(
            header_frame,
            text="NOTICE",
            font=("Helvetica", 24, "bold"),
            fg="white",
            bg="#4b7bec"
        )
        title_label.pack(pady=10)

        # Информация о API ключах
        api_frame = ttk.LabelFrame(self.root, text="API ключи")
        api_frame.pack(fill=tk.X, padx=10, pady=10)

        api_info_label = ttk.Label(
            api_frame,
            text=f"Загружено {len(self.api_keys)} API ключей для работы с Checko API",
            font=("Helvetica", 10)
        )
        api_info_label.pack(padx=5, pady=5)

        # Панель параметров поиска
        params_frame = ttk.LabelFrame(self.root, text="Параметры поиска")
        params_frame.pack(fill=tk.X, padx=10, pady=5)

        # ОКВЭД
        ttk.Label(params_frame, text="Код ОКВЭД:").grid(row=0, column=0, padx=5, pady=10, sticky="w")
        self.okved_entry = ttk.Entry(params_frame, width=10)
        self.okved_entry.grid(row=0, column=1, padx=5, pady=10, sticky="w")

        # Максимальное количество записей
        ttk.Label(params_frame, text="Макс. количество:").grid(row=0, column=2, padx=5, pady=10, sticky="w")
        self.max_entries_entry = ttk.Entry(params_frame, width=7)
        self.max_entries_entry.insert(0, "1000")
        self.max_entries_entry.grid(row=0, column=3, padx=5, pady=10, sticky="w")

        # Только активные
        self.active_only_var = tk.BooleanVar(value=True)
        active_check = ttk.Checkbutton(params_frame, text="Только активные организации", variable=self.active_only_var)
        active_check.grid(row=0, column=4, padx=5, pady=10, sticky="w")

        # Кнопка поиска
        search_btn = ttk.Button(params_frame, text="Начать поиск", command=self.start_search)
        search_btn.grid(row=0, column=5, padx=20, pady=10)

        self.stop_btn = ttk.Button(params_frame, text="Остановить", command=self.stop_search, state=tk.DISABLED)
        self.stop_btn.grid(row=0, column=6, padx=5, pady=10)

        # Панель результатов
        results_frame = ttk.LabelFrame(self.root, text="Результаты")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Таблица с результатами
        columns = ("inn", "name", "address", "email", "phone")
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show="headings")

        # Настройка столбцов
        self.results_tree.heading("inn", text="ИНН")
        self.results_tree.heading("name", text="Наименование")
        self.results_tree.heading("address", text="Адрес")
        self.results_tree.heading("email", text="Email")
        self.results_tree.heading("phone", text="Телефон")

        self.results_tree.column("inn", width=100)
        self.results_tree.column("name", width=250)
        self.results_tree.column("address", width=300)
        self.results_tree.column("email", width=150)
        self.results_tree.column("phone", width=150)

        # Полосы прокрутки
        y_scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.results_tree.yview)
        x_scrollbar = ttk.Scrollbar(results_frame, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)

        # Размещение элементов
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Кнопки действий под результатами
        actions_frame = ttk.Frame(self.root)
        actions_frame.pack(fill=tk.X, padx=10, pady=5)

        export_btn = ttk.Button(actions_frame, text="Экспорт в Excel", command=self.export_to_xlsx)
        export_btn.pack(side=tk.RIGHT, padx=5)

        # Статус бар
        status_frame = tk.Frame(self.root, height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.status_var = tk.StringVar(value="Готов к работе")
        status_label = ttk.Label(status_frame, textvariable=self.status_var)
        status_label.pack(side=tk.LEFT, padx=10)

        # Прогресс бар
        self.progress_var = tk.DoubleVar(value=0)
        self.progress = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress.pack(fill=tk.X, padx=10, pady=5, side=tk.BOTTOM)

        # Канвас для анимаций
        self.canvas = tk.Canvas(self.root, height=5, bg="#f0f0f0", highlightthickness=0)
        self.canvas.pack(fill=tk.X, side=tk.BOTTOM)

    def animate(self):
        """Анимация индикатора во время поиска"""
        colors = ["#4b7bec", "#45aaf2", "#2ecc71", "#fed330", "#fd9644"]
        width = self.root.winfo_width()
        x = 0

        while self.animation_running:
            color = colors[x % len(colors)]
            obj = self.canvas.create_rectangle(x, 0, x + 20, 5, fill=color, outline="")
            self.animation_objects.append(obj)

            # Удаление старых объектов для предотвращения утечки памяти
            if len(self.animation_objects) > 50:
                old_obj = self.animation_objects.pop(0)
                self.canvas.delete(old_obj)

            x = (x + 20) % width
            self.root.update()
            time.sleep(0.03)

    def start_animation(self):
        """Запуск анимации"""
        self.animation_running = True
        self.canvas.delete("all")  # Очистка канваса
        threading.Thread(target=self.animate, daemon=True).start()

    def stop_animation(self):
        """Остановка анимации"""
        self.animation_running = False
        self.canvas.delete("all")

    def make_request_with_key_rotation(self, url, params, api_key_manager, request_name=""):
        """Выполнить запрос с автоматической ротацией ключей при ошибках"""
        # Пробуем все ключи, пока один не сработает или не закончатся
        max_attempts = len(api_key_manager.api_keys)
        attempts = 0

        while attempts < max_attempts and self.search_running:
            params["key"] = api_key_manager.get_current_key()

            try:
                if request_name:
                    self.status_var.set(f"{request_name} (ключ #{api_key_manager.current_index + 1})")

                response = requests.get(url, params=params, timeout=10)
                data = response.json()

                # Проверка на наличие ошибок в ответе
                if "error" in data.get("meta", {}).get("status", "").lower():
                    error_message = data.get("meta", {}).get("message", "Неизвестная ошибка")
                    self.status_var.set(f"Ошибка ключа #{api_key_manager.current_index + 1}: {error_message}")

                    # Пробуем следующий ключ
                    api_key_manager.rotate_key()
                    attempts += 1
                    time.sleep(0.5)  # Небольшая задержка перед следующей попыткой
                    continue

                # Если дошли до этой точки, значит запрос выполнен успешно
                return data

            except Exception as e:
                self.status_var.set(f"Ошибка запроса с ключом #{api_key_manager.current_index + 1}: {str(e)}")
                api_key_manager.rotate_key()
                attempts += 1
                time.sleep(0.5)

        # Если все ключи испробованы и ни один не сработал
        if attempts >= max_attempts:
            self.status_var.set("Все API ключи недействительны или исчерпан лимит запросов")
            return None

        return None

    def search_companies(self, api_key_manager, okved, max_entries, active_only):
        """Поиск организаций по ОКВЭД с поддержкой пагинации и дедупликацией"""
        all_companies = []
        seen_inns = set()  # Для отслеживания уже полученных ИНН
        page = 1
        limit = 100  # API Checko возвращает максимум 100 записей за запрос
        total_found = 0
        duplicate_pages_count = 0
        max_duplicate_pages = 3  # Максимальное количество страниц без новых данных

        while self.search_running:
            url = "https://api.checko.ru/v2/search"
            params = {
                "by": "okved",
                "obj": "org",
                "query": okved,
                "limit": limit,
                "page": page
            }

            if active_only:
                params["active"] = "true"

            self.status_var.set(f"Загрузка страницы {page}...")

            # Выполняем запрос с автоматической ротацией ключей
            data = self.make_request_with_key_rotation(
                url,
                params,
                api_key_manager,
                f"Поиск организаций (стр. {page})"
            )

            if not data:
                break

            if "data" in data and "Записи" in data["data"]:
                companies = data["data"]["Записи"]
                if not companies:
                    self.status_var.set("Больше записей не найдено")
                    break

                # Подсчитываем новые компании на этой странице
                new_companies_count = 0
                for company in companies:
                    inn = company.get("ИНН", "")
                    if inn and inn not in seen_inns:
                        seen_inns.add(inn)
                        all_companies.append(company)
                        new_companies_count += 1

                        # Если достигли максимального количества записей
                        if len(all_companies) >= max_entries:
                            self.status_var.set(f"Достигнут лимит в {max_entries} компаний")
                            break

                # Если на текущей странице нет новых компаний, увеличиваем счетчик дубликатов
                if new_companies_count == 0:
                    duplicate_pages_count += 1
                    self.status_var.set(
                        f"Страница {page} не содержит новых данных. Дубликатов: {duplicate_pages_count}")

                    # Если получили несколько страниц без новых данных, заканчиваем
                    if duplicate_pages_count >= max_duplicate_pages:
                        self.status_var.set(f"Получено максимальное количество данных ({len(all_companies)} компаний)")
                        break
                else:
                    # Сбрасываем счетчик, если нашли новые компании
                    duplicate_pages_count = 0

                # Получаем общее количество найденных записей из мета-данных
                if "meta" in data and "found" in data["meta"]:
                    total_found = data["meta"]["found"]
                    self.status_var.set(f"Найдено всего: {total_found}, получено уникальных: {len(all_companies)}")

                # Если достигли максимума или собрали все уникальные записи
                if len(all_companies) >= max_entries or (total_found > 0 and len(all_companies) >= total_found):
                    break

                # Переход к следующей странице
                page += 1
            else:
                # Нет данных в ответе API
                self.status_var.set("Нет данных по запросу")
                break

        return all_companies

    def get_company_details(self, api_key_manager, inn):
        """Получение детальной информации о компании по ИНН"""
        url = "https://api.checko.ru/v2/company"
        params = {"inn": inn}

        # Выполняем запрос с автоматической ротацией ключей
        data = self.make_request_with_key_rotation(
            url,
            params,
            api_key_manager,
            f"Получение данных ИНН {inn}"
        )

        if not data or "data" not in data:
            return None

        company_data = data["data"]

        # Извлечение нужных данных - исправлено получение наименования
        result = {
            "inn": inn,
            "name": "",
            "address": "",
            "email": "",
            "phone": ""
        }

        # Получение наименования - проверяем разные варианты полей
        if "НаимСокр" in company_data:
            result["name"] = company_data["НаимСокр"]
        elif "НаимПолн" in company_data:
            result["name"] = company_data["НаимПолн"]
        elif "Наим Сокр" in company_data:
            result["name"] = company_data["Наим Сокр"]
        elif "Наим Полн" in company_data:
            result["name"] = company_data["Наим Полн"]

        # Получение адреса
        if "ЮрАдрес" in company_data:
            if isinstance(company_data["ЮрАдрес"], dict):
                if "АдресПолн" in company_data["ЮрАдрес"]:
                    result["address"] = company_data["ЮрАдрес"]["АдресПолн"]
                elif "АдресРФ" in company_data["ЮрАдрес"]:
                    result["address"] = company_data["ЮрАдрес"]["АдресРФ"]
            elif isinstance(company_data["ЮрАдрес"], str):
                result["address"] = company_data["ЮрАдрес"]

        # Получение контактных данных
        if "Контакты" in company_data:
            contacts = company_data["Контакты"]

            # Email
            if "Емэйл" in contacts and contacts["Емэйл"] and len(contacts["Емэйл"]) > 0:
                result["email"] = contacts["Емэйл"][0]

            # Телефон
            if "Тел" in contacts and contacts["Тел"] and len(contacts["Тел"]) > 0:
                result["phone"] = contacts["Тел"][0]

        return result

    def process_search(self):
        """Основная функция поиска"""
        # Получение параметров
        okved = self.okved_entry.get().strip()
        if not okved:
            messagebox.showerror("Ошибка", "Введите код ОКВЭД")
            self.stop_btn.config(state=tk.DISABLED)
            return

        try:
            max_entries = int(self.max_entries_entry.get().strip())
            if max_entries <= 0:
                raise ValueError("Максимальное количество должно быть положительным числом")
        except ValueError:
            messagebox.showerror("Ошибка", "Максимальное количество должно быть положительным числом")
            self.stop_btn.config(state=tk.DISABLED)
            return

        active_only = self.active_only_var.get()

        # Очистка предыдущих результатов
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)

        self.collected_data = []

        # Настройка UI для поиска
        self.stop_btn.config(state=tk.NORMAL)
        self.start_animation()
        self.status_var.set(f"Поиск организаций с ОКВЭД: {okved}...")
        self.progress_var.set(0)

        # Создание менеджера API ключей
        api_manager = APIKeyManager(self.api_keys)

        # Поиск организаций
        companies = self.search_companies(api_manager, okved, max_entries, active_only)

        if not companies:
            self.status_var.set("Организации не найдены или все API ключи недействительны")
            self.stop_animation()
            self.stop_btn.config(state=tk.DISABLED)
            return

        total = len(companies)
        self.status_var.set(f"Найдено {total} уникальных организаций. Получение детальной информации...")

        # Получение детальной информации по каждой организации
        for i, company in enumerate(companies):
            if not self.search_running:
                break

            inn = company.get("ИНН", "")
            if not inn:
                continue

            # Обновление статуса и прогресса
            self.status_var.set(f"Обработка {i + 1} из {total}: ИНН {inn}")
            self.progress_var.set((i / total) * 100)
            self.root.update()  # Обновляем UI чтобы показать прогресс

            # Получение детальной информации
            details = self.get_company_details(api_manager, inn)
            if details:
                self.collected_data.append(details)

                # Добавление в таблицу
                self.results_tree.insert("", tk.END, values=(
                    details["inn"],
                    details["name"],
                    details["address"],
                    details["email"],
                    details["phone"]
                ))
                self.root.update()  # Обновляем UI после добавления строки

            # Небольшая задержка, чтобы не перегружать API
            time.sleep(0.3)

        # Завершение поиска
        self.status_var.set(f"Поиск завершен. Найдено {len(self.collected_data)} организаций с данными.")
        self.progress_var.set(100)
        self.stop_animation()
        self.stop_btn.config(state=tk.DISABLED)

    def start_search(self):
        """Запуск поиска в отдельном потоке"""
        self.search_running = True
        threading.Thread(target=self.process_search, daemon=True).start()

    def stop_search(self):
        """Остановка поиска"""
        self.search_running = False
        self.status_var.set("Поиск остановлен пользователем")

    def export_to_xlsx(self):
        """Экспорт результатов в Excel файл"""
        if not self.collected_data:
            messagebox.showinfo("Информация", "Нет данных для экспорта")
            return

        # Формирование имени файла с текущей датой и временем
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"notice_export_{now}.xlsx"

        try:
            # Создаем новую книгу Excel и активный лист
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Организации"

            # Заголовки столбцов
            headers = ["ИНН", "Наименование", "Адрес", "Email", "Телефон"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Установка ширины столбцов
            ws.column_dimensions['A'].width = 15  # ИНН
            ws.column_dimensions['B'].width = 40  # Наименование
            ws.column_dimensions['C'].width = 50  # Адрес
            ws.column_dimensions['D'].width = 25  # Email
            ws.column_dimensions['E'].width = 20  # Телефон

            # Заполнение данными
            for row_num, data in enumerate(self.collected_data, 2):
                ws.cell(row=row_num, column=1, value=data["inn"])
                ws.cell(row=row_num, column=2, value=data["name"])
                ws.cell(row=row_num, column=3, value=data["address"])
                ws.cell(row=row_num, column=4, value=data["email"])
                ws.cell(row=row_num, column=5, value=data["phone"])

            # Добавление границ для всех заполненных ячеек
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=len(self.collected_data) + 1, max_col=5):
                for cell in row:
                    cell.border = thin_border

            # Сохранение книги
            wb.save(filename)
            messagebox.showinfo("Успешно", f"Данные экспортированы в файл: {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте: {str(e)}")


# Запуск программы
def main():
    root = tk.Tk()
    app = NoticeApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
